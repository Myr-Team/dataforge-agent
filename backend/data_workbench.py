from __future__ import annotations

import base64
import csv
import hashlib
import io
import json
import os
import re
import time
import uuid
from collections import Counter
from contextlib import closing
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from fastapi import APIRouter, BackgroundTasks, HTTPException, Query, Request
from openpyxl import Workbook, load_workbook
from starlette.concurrency import run_in_threadpool

from ingest.profiler import build_data_profile, profile_to_search_document
from ingest.adapters.upload_to_records import upload_to_records

try:
    from .blob_store import delete_blob_name, download_blob_content, upload_workspace_blob
    from .identity import actor_for_history, actor_from_request, merge_actor_into_ui_context
    from .orchestrator import orchestrate_chat
    from .schemas import ChatRequest
    from .workspace_authz import require_workspace_permission
    from .workspace_store import (
        WORKSPACES,
        _CONTEXT_CACHE,
        _index_documents_batched,
        _load_workspace_bundle,
        _normalize_documents,
        _persist_workspace_state,
        _rebuild_workspace_profile,
        create_workspace_upload_job,
        get_workspace_detail,
        run_workspace_ingest_job,
    )
except ImportError:
    from blob_store import delete_blob_name, download_blob_content, upload_workspace_blob
    from identity import actor_for_history, actor_from_request, merge_actor_into_ui_context
    from orchestrator import orchestrate_chat
    from schemas import ChatRequest
    from workspace_authz import require_workspace_permission
    from workspace_store import (
        WORKSPACES,
        _CONTEXT_CACHE,
        _index_documents_batched,
        _load_workspace_bundle,
        _normalize_documents,
        _persist_workspace_state,
        _rebuild_workspace_profile,
        create_workspace_upload_job,
        get_workspace_detail,
        run_workspace_ingest_job,
    )


router = APIRouter(prefix="/api/workspaces/{workspace_id}", tags=["data-workbench"])

MAX_PREVIEW_LIMIT = int(os.environ.get("DF_WORKBENCH_PREVIEW_LIMIT", "500"))
MAX_MARKDOWN_BYTES = int(os.environ.get("DF_WORKBENCH_MARKDOWN_SAVE_BYTES", str(512 * 1024)))
MAX_TABLE_SAVE_BYTES = int(os.environ.get("DF_WORKBENCH_TABLE_SAVE_BYTES", str(4 * 1024 * 1024)))
MAX_CELL_EDITS = int(os.environ.get("DF_WORKBENCH_MAX_CELL_EDITS", "500"))
MAX_CELL_CHARS = int(os.environ.get("DF_WORKBENCH_MAX_CELL_CHARS", "5000"))
MAX_CONNECTOR_IMPORT_BYTES = int(os.environ.get("DF_CONNECTOR_IMPORT_MAX_BYTES", str(25 * 1024 * 1024)))
CONNECTOR_TTL_SECONDS = int(os.environ.get("DF_CONNECTOR_SESSION_TTL_SECONDS", "3600"))
STORAGE_TOTAL_BYTES = int(os.environ.get("DF_WORKSPACE_STORAGE_TOTAL_BYTES", str(5 * 1024 * 1024 * 1024)))

TABLE_TYPES = {"csv", "xlsx", "xlsm", "excel"}
MARKDOWN_TYPES = {"md", "markdown", "txt", "text"}
JSON_TYPES = {"json"}


@router.get("/files")
async def workspace_files(workspace_id: str, request: Request) -> dict[str, Any]:
    _require(request, workspace_id, "file.read")
    return await _call(list_workspace_files, workspace_id)


@router.post("/files")
async def workspace_file_create(workspace_id: str, request: Request) -> dict[str, Any]:
    _require(request, workspace_id, "file.create")
    body = await _json_body(request)
    return await _call(create_workspace_file, workspace_id, body, actor_from_request(request))


@router.get("/files/{file_id}/content")
async def workspace_file_content(
    workspace_id: str,
    file_id: str,
    request: Request,
    limit: int = Query(default=100, ge=1, le=MAX_PREVIEW_LIMIT),
    offset: int = Query(default=0, ge=0),
) -> dict[str, Any]:
    _require(request, workspace_id, "file.read")
    return await _call(preview_file_content, workspace_id, file_id, limit, offset)


@router.put("/files/{file_id}/cells")
async def workspace_file_cells(workspace_id: str, file_id: str, request: Request) -> dict[str, Any]:
    _require(request, workspace_id, "file.edit")
    body = await _json_body(request)
    return await _call(save_table_cells, workspace_id, file_id, body, actor_from_request(request))


@router.put("/files/{file_id}/content")
async def workspace_file_markdown(workspace_id: str, file_id: str, request: Request) -> dict[str, Any]:
    _require(request, workspace_id, "file.edit")
    body = await _json_body(request)
    return await _call(save_markdown_content, workspace_id, file_id, body, actor_from_request(request))


@router.delete("/files/{file_id}")
async def workspace_file_delete(workspace_id: str, file_id: str, request: Request) -> dict[str, Any]:
    _require(request, workspace_id, "file.delete")
    return await _call(delete_workspace_file, workspace_id, file_id, actor_from_request(request))


@router.get("/files/{file_id}/quality")
async def workspace_file_quality(workspace_id: str, file_id: str, request: Request) -> dict[str, Any]:
    _require(request, workspace_id, "file.read")
    return await _call(file_quality, workspace_id, file_id)


@router.get("/files/{file_id}/field-mapping")
async def workspace_file_field_mapping(workspace_id: str, file_id: str, request: Request) -> dict[str, Any]:
    _require(request, workspace_id, "file.read")
    return await _call(file_field_mapping, workspace_id, file_id)


@router.put("/files/{file_id}/field-mapping")
async def workspace_file_field_mapping_save(workspace_id: str, file_id: str, request: Request) -> dict[str, Any]:
    _require(request, workspace_id, "file.edit")
    body = await _json_body(request)
    return await _call(save_file_field_mapping, workspace_id, file_id, body, actor_from_request(request))


@router.get("/files/{file_id}/history")
async def workspace_file_history(workspace_id: str, file_id: str, request: Request) -> list[dict[str, Any]]:
    _require(request, workspace_id, "file.read")
    return await _call(file_history, workspace_id, file_id)


@router.post("/files/analyze")
async def workspace_files_analyze(workspace_id: str, request: Request, background_tasks: BackgroundTasks) -> dict[str, Any]:
    _require(request, workspace_id, "analysis.run")
    body = await _json_body(request)
    body["ui_context"] = merge_actor_into_ui_context(body.get("ui_context") if isinstance(body.get("ui_context"), dict) else {}, actor_from_request(request))
    return await analyze_selected_files(workspace_id, body, background_tasks)


@router.get("/connectors/capabilities")
async def connector_capabilities(workspace_id: str, request: Request) -> dict[str, Any]:
    _require(request, workspace_id, "workspace.read")
    return {
        "workspace_id": workspace_id,
        "connectors": [
            {"kind": "azure_data_lake", "status": "demo", "mode": "demo_only"},
            {"kind": "azure_blob", "status": "available", "mode": "manual_credentials"},
            {"kind": "sql_database", "status": "available", "mode": "manual_credentials"},
            {
                "kind": "identity_discovery",
                "status": "not_configured",
                "mode": "placeholder",
                "requirements": ["Easy Auth token store", "ARM/Storage/SQL delegated permissions", "target resource RBAC"],
            },
        ],
    }


@router.post("/connectors/sql/connect")
async def sql_connect(workspace_id: str, request: Request) -> dict[str, Any]:
    _require(request, workspace_id, "connector.manage")
    body = await _json_body(request)
    return await _call(connect_sql, workspace_id, body)


@router.get("/connectors/sql/status")
async def sql_status(workspace_id: str, connection_id: str, request: Request) -> dict[str, Any]:
    _require(request, workspace_id, "workspace.read")
    return await _call(connector_status, workspace_id, "sql", connection_id)


@router.get("/connectors/sql/tables")
async def sql_tables(workspace_id: str, connection_id: str, request: Request) -> dict[str, Any]:
    _require(request, workspace_id, "workspace.read")
    return await _call(list_sql_tables, workspace_id, connection_id)


@router.get("/connectors/sql/preview")
async def sql_preview(workspace_id: str, connection_id: str, table: str, request: Request, limit: int = Query(default=100, ge=1, le=1000)) -> dict[str, Any]:
    _require(request, workspace_id, "workspace.read")
    return await _call(preview_sql_table, workspace_id, connection_id, table, limit)


@router.post("/connectors/sql/import")
async def sql_import(workspace_id: str, request: Request) -> dict[str, Any]:
    _require(request, workspace_id, "connector.manage")
    body = await _json_body(request)
    return await _call(import_sql_table, workspace_id, body, actor_from_request(request))


@router.post("/connectors/blob/connect")
async def blob_connect(workspace_id: str, request: Request) -> dict[str, Any]:
    _require(request, workspace_id, "connector.manage")
    body = await _json_body(request)
    return await _call(connect_blob, workspace_id, body)


@router.get("/connectors/blob/status")
async def blob_status(workspace_id: str, connection_id: str, request: Request) -> dict[str, Any]:
    _require(request, workspace_id, "workspace.read")
    return await _call(connector_status, workspace_id, "blob", connection_id)


@router.get("/connectors/blob/containers")
async def blob_containers(workspace_id: str, connection_id: str, request: Request) -> dict[str, Any]:
    _require(request, workspace_id, "workspace.read")
    return await _call(list_blob_containers, workspace_id, connection_id)


@router.get("/connectors/blob/blobs")
async def blob_blobs(
    workspace_id: str,
    connection_id: str,
    container: str,
    request: Request,
    prefix: str = "",
    limit: int = Query(default=100, ge=1, le=500),
) -> dict[str, Any]:
    _require(request, workspace_id, "workspace.read")
    return await _call(list_blob_items, workspace_id, connection_id, container, prefix, limit)


@router.get("/connectors/blob/preview")
async def blob_preview(
    workspace_id: str,
    connection_id: str,
    container: str,
    blob: str,
    request: Request,
    limit: int = Query(default=100, ge=1, le=MAX_PREVIEW_LIMIT),
    offset: int = Query(default=0, ge=0),
) -> dict[str, Any]:
    _require(request, workspace_id, "workspace.read")
    return await _call(preview_blob_item, workspace_id, connection_id, container, blob, limit, offset)


@router.post("/connectors/blob/import")
async def blob_import(workspace_id: str, request: Request) -> dict[str, Any]:
    _require(request, workspace_id, "connector.manage")
    body = await _json_body(request)
    return await _call(import_blob_item, workspace_id, body, actor_from_request(request))


@router.post("/connectors/disconnect")
async def connectors_disconnect(workspace_id: str, request: Request) -> dict[str, Any]:
    _require(request, workspace_id, "connector.manage")
    body = await _json_body(request)
    return await _call(disconnect_connector, workspace_id, body)


def list_workspace_files(workspace_id: str) -> dict[str, Any]:
    detail = get_workspace_detail(workspace_id)
    documents = [item for item in detail.get("documents") or [] if isinstance(item, dict)]
    field_counts = _document_field_counts(detail, documents)
    single_doc_fields = int(detail.get("field_count") or 0) if len(documents) == 1 else 0
    groups = [
        {"label": "数据集", "files": []},
        {"label": "文档", "files": []},
        {"label": "临时文件", "files": []},
    ]
    by_label = {item["label"]: item["files"] for item in groups}
    used_bytes = 0
    for document in documents:
        source = str(document.get("source_file") or "")
        entry = _file_entry(document, field_count=field_counts.get(source) or single_doc_fields)
        _fill_file_counts_from_content(workspace_id, document, entry)
        used_bytes += int(entry.get("bytes") or 0)
        label = _file_group(entry)
        by_label[label].append(entry)
    return {
        "workspace_id": workspace_id,
        "groups": groups,
        "storage": {"used_bytes": used_bytes, "total_bytes": STORAGE_TOTAL_BYTES},
    }


def preview_file_content(workspace_id: str, file_id: str, limit: int = 100, offset: int = 0) -> dict[str, Any]:
    document = _find_document(workspace_id, file_id)
    content, _content_type = _read_document_bytes(workspace_id, document)
    return _preview_bytes(content, _file_type(document), Path(str(document.get("name") or "")).name, limit, offset)


def create_workspace_file(workspace_id: str, body: dict[str, Any], actor: dict[str, Any] | None = None) -> dict[str, Any]:
    filename, file_type = _new_file_name_and_type(body)
    content, content_type = _new_file_content(body, file_type)
    result = create_workspace_upload_job(
        files=[{"filename": filename, "content": content, "content_type": content_type}],
        name=None,
        requested_workspace_id=workspace_id,
        asset_role=str(body.get("asset_role") or "reference"),
    )
    created_sources = {
        str(item.get("source_file") or "")
        for item in result.get("documents") or []
        if isinstance(item, dict) and item.get("ingest_job_id") == result.get("ingest_job_id")
    }
    _run_ingest_if_present(result)
    created = None
    files = list_workspace_files(workspace_id)
    for group in files.get("groups") or []:
        for item in group.get("files") or []:
            if isinstance(item, dict) and (str(item.get("source_file") or "") in created_sources or item.get("name") == filename):
                created = item
                break
        if created:
            break
    if created:
        _append_workbench_history(
            workspace_id,
            str(created.get("id") or ""),
            actor,
            _utc_now(),
            f"Created {created.get('name') or filename}",
        )
    return {
        "workspace_id": workspace_id,
        "saved_at": _utc_now(),
        "file": created or {"name": filename, "type": file_type},
        "upload": _compact_upload_result(result),
    }


def save_table_cells(workspace_id: str, file_id: str, body: dict[str, Any], actor: dict[str, Any] | None = None) -> dict[str, Any]:
    document = _find_document(workspace_id, file_id)
    file_type = _file_type(document)
    if file_type not in TABLE_TYPES:
        raise ValueError("Only CSV/XLSX files support cell edits")
    edits = body.get("edits") or []
    operations = _table_operations_from_body(body)
    if not isinstance(edits, list):
        raise ValueError("edits must be a list")
    if not edits and not operations:
        raise ValueError("edits or table operations are required")
    if len(edits) > MAX_CELL_EDITS:
        raise ValueError(f"Too many edits; max {MAX_CELL_EDITS}")
    content, _ = _read_document_bytes(workspace_id, document)
    if len(content) > MAX_TABLE_SAVE_BYTES:
        raise ValueError("File is too large for lightweight cell editing")

    version = _snapshot_version(workspace_id, document, content)
    if file_type == "csv":
        new_content, row_count, col_count, change_counts = _apply_csv_table_update(content, edits, operations)
        content_type = "text/csv; charset=utf-8"
    else:
        new_content, row_count, col_count, change_counts = _apply_xlsx_table_update(content, edits, operations)
        content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    saved_at = _utc_now()
    summary = _table_change_summary(change_counts)
    _write_document_bytes(workspace_id, document, new_content, content_type)
    _update_document_after_save(workspace_id, document, new_content, saved_at, summary, version, row_count=row_count, col_count=col_count, actor=actor)
    return {
        "saved_at": saved_at,
        "version_id": version["version_id"],
        "change_summary": summary,
        "row_count": row_count,
        "col_count": col_count,
        "changes": change_counts,
    }


def save_markdown_content(workspace_id: str, file_id: str, body: dict[str, Any], actor: dict[str, Any] | None = None) -> dict[str, Any]:
    document = _find_document(workspace_id, file_id)
    file_type = _file_type(document)
    if file_type not in MARKDOWN_TYPES:
        raise ValueError("Only markdown/text files support content edits")
    text = body.get("text")
    if not isinstance(text, str):
        raise ValueError("text is required")
    content = text.encode("utf-8")
    if len(content) > MAX_MARKDOWN_BYTES:
        raise ValueError("Markdown content is too large")
    previous, _ = _read_document_bytes(workspace_id, document)
    version = _snapshot_version(workspace_id, document, previous)
    saved_at = _utc_now()
    summary = "更新了 Markdown 内容"
    _write_document_bytes(workspace_id, document, content, "text/markdown; charset=utf-8")
    line_count = len(text.splitlines())
    _update_document_after_save(workspace_id, document, content, saved_at, summary, version, row_count=line_count, col_count=1, actor=actor)
    return {"saved_at": saved_at, "version_id": version["version_id"], "change_summary": summary}


def delete_workspace_file(workspace_id: str, file_id: str, actor: dict[str, Any] | None = None) -> dict[str, Any]:
    meta, profile = _workspace_state(workspace_id)
    documents = meta.get("documents")
    if not isinstance(documents, list):
        documents = _normalize_documents(WORKSPACES / workspace_id, meta, profile)
        meta["documents"] = documents
    document = _find_document(workspace_id, file_id, detail={"documents": documents})
    fid = _file_id(document)
    source = _safe_source_file(document)
    profile_file = str(document.get("profile_file") or "")
    removed_at = _utc_now()

    meta["documents"] = [
        item for item in documents if not (isinstance(item, dict) and str(item.get("source_file") or "") == source)
    ]
    if isinstance(meta.get("workbench_field_mappings"), dict):
        meta["workbench_field_mappings"].pop(fid, None)
    deleted = meta.setdefault("workbench_deleted_files", [])
    if isinstance(deleted, list):
        deleted.insert(
            0,
            {
                "file_id": fid,
                "name": document.get("name") or Path(source).name,
                "source_file": source,
                "profile_file": profile_file or None,
                "deleted_at": removed_at,
                "actor": actor_for_history(actor),
            },
        )
        del deleted[50:]
    if isinstance(meta.get("workbench_history"), dict):
        meta["workbench_history"].pop(fid, None)

    for rel in [source, profile_file]:
        if not rel:
            continue
        local_path = WORKSPACES / workspace_id / rel
        try:
            if local_path.exists() and local_path.is_file():
                local_path.unlink()
        except OSError:
            pass
        try:
            delete_blob_name(f"workspaces/{workspace_id}/{rel}")
        except Exception:
            pass

    profile = _rebuild_workspace_profile(workspace_id, WORKSPACES / workspace_id, meta, profile)
    _persist_workspace_state(workspace_id, WORKSPACES / workspace_id, meta, profile, include_raw_payloads=True)
    _CONTEXT_CACHE.pop(workspace_id, None)
    return {
        "workspace_id": workspace_id,
        "file_id": fid,
        "deleted": True,
        "deleted_at": removed_at,
        "name": document.get("name") or Path(source).name,
    }


def file_quality(workspace_id: str, file_id: str) -> dict[str, Any]:
    document = _find_document(workspace_id, file_id)
    content, _content_type = _read_document_bytes(workspace_id, document)
    file_type = _file_type(document)
    checked_at = _utc_now()
    if file_type in MARKDOWN_TYPES:
        text = _decode_text(content)
        mapped = 1 if text.strip() else 0
        status = "passed" if mapped else "failed"
        result = {
            "workspace_id": workspace_id,
            "file_id": _file_id(document),
            "field_mapping": {"mapped": mapped, "total": 1, "pct": 100.0 if mapped else 0.0, "fields": [{"name": "text", "type": "markdown", "mapped": bool(mapped)}]},
            "quality": {
                "missing_pct": 0.0 if mapped else 100.0,
                "duplicate_pct": 0.0,
                "outlier_count": 0,
                "outlier_pct": 0.0,
                "type_warnings": 0,
                "row_count": len(text.splitlines()),
                "column_count": 1,
            },
            "validation": {"status": status, "checked_at": checked_at},
        }
        return _apply_mapping_overrides(workspace_id, _file_id(document), result)
    if file_type in JSON_TYPES:
        return _apply_mapping_overrides(workspace_id, _file_id(document), _quality_from_json(workspace_id, document, content, checked_at))
    table = _table_from_bytes(content, file_type)
    return _apply_mapping_overrides(workspace_id, _file_id(document), _quality_from_table(workspace_id, document, table, checked_at))


def file_field_mapping(workspace_id: str, file_id: str) -> dict[str, Any]:
    quality = file_quality(workspace_id, file_id)
    meta, _profile = _workspace_state(workspace_id)
    fid = str(quality.get("file_id") or file_id)
    stored = _stored_field_mapping(meta, fid)
    return {
        "workspace_id": workspace_id,
        "file_id": fid,
        "field_mapping": quality.get("field_mapping") or {},
        "overrides": stored.get("mappings") or {},
        "saved_at": stored.get("saved_at"),
        "validation": quality.get("validation") or {},
    }


def save_file_field_mapping(workspace_id: str, file_id: str, body: dict[str, Any], actor: dict[str, Any] | None = None) -> dict[str, Any]:
    document = _find_document(workspace_id, file_id)
    fid = _file_id(document)
    quality = file_quality(workspace_id, fid)
    fields = {
        str(item.get("name") or "")
        for item in ((quality.get("field_mapping") or {}).get("fields") or [])
        if isinstance(item, dict) and item.get("name")
    }
    mappings = _normalize_field_mapping_body(body, fields)
    saved_at = _utc_now()
    meta, profile = _workspace_state(workspace_id)
    all_mappings = meta.setdefault("workbench_field_mappings", {})
    all_mappings[fid] = {
        "file_id": fid,
        "source_file": document.get("source_file"),
        "mappings": mappings,
        "saved_at": saved_at,
    }
    _persist_workspace_state(workspace_id, WORKSPACES / workspace_id, meta, profile, include_raw_payloads=True)
    _append_workbench_history(workspace_id, fid, actor, saved_at, "Updated field mapping")
    _CONTEXT_CACHE.pop(workspace_id, None)
    updated = file_field_mapping(workspace_id, fid)
    updated["saved_at"] = saved_at
    return updated


def file_history(workspace_id: str, file_id: str) -> list[dict[str, Any]]:
    meta, _profile = _workspace_state(workspace_id)
    document = _find_document(workspace_id, file_id, detail=get_workspace_detail(workspace_id))
    fid = _file_id(document)
    history = (meta.get("workbench_history") or {}).get(fid) if isinstance(meta.get("workbench_history"), dict) else None
    if isinstance(history, list) and history:
        return [item for item in history if isinstance(item, dict)]
    at = document.get("updated_at") or document.get("created_at") or meta.get("updated_at") or meta.get("created_at") or _utc_now()
    return [{"user": "DataForge", "email": None, "at": at, "change_summary": "初始文件版本"}]


async def analyze_selected_files(workspace_id: str, body: dict[str, Any], background_tasks: BackgroundTasks | None = None) -> dict[str, Any]:
    raw_ids = body.get("file_ids") or body.get("files") or []
    file_ids = [str(item) for item in raw_ids if str(item).strip()]
    if not file_ids:
        raise HTTPException(status_code=400, detail="file_ids is required")
    detail = get_workspace_detail(workspace_id)
    selected_documents = [_find_document(workspace_id, file_id, detail=detail) for file_id in file_ids]
    warnings: list[str] = []
    for document in selected_documents:
        if _api_status(document) == "indexed":
            continue
        job_id = str(document.get("ingest_job_id") or "").strip()
        if not job_id:
            warnings.append(f"{document.get('name') or _file_id(document)} 尚未完成索引，但文件内容可直接引用。")
            continue
        try:
            run_workspace_ingest_job(workspace_id, job_id)
            warnings.append(f"{document.get('name') or _file_id(document)} 已触发补索引。")
        except Exception as exc:
            warnings.append(f"{document.get('name') or _file_id(document)} 补索引未完成：{type(exc).__name__}")
    detail = get_workspace_detail(workspace_id)
    selected_documents = [_find_document(workspace_id, file_id, detail=detail) for file_id in file_ids]
    unavailable: list[str] = []
    for document in selected_documents:
        try:
            _read_document_bytes(workspace_id, document)
        except Exception:
            unavailable.append(str(document.get("name") or _file_id(document)))
    if unavailable:
        raise HTTPException(
            status_code=409,
            detail={
                "message": "选中的文件还不能用于分析，请等待上传/索引完成或重新选择文件。",
                "unavailable_files": unavailable,
                "recoverable": True,
            },
        )
    selected = [_file_entry(document) for document in selected_documents]
    names = ", ".join(item["name"] for item in selected[:6])
    message = str(body.get("message") or "").strip() or (
        f"请只基于数据工作台中选中的文件（{names}）发起一次数据商机化分析，输出机会、证据强度、风险缺口和下一步验证计划。"
    )
    conversation_id = str(body.get("conversation_id") or "").strip() or str(uuid.uuid4())
    req = ChatRequest(
        workspace_id=workspace_id,
        message=message,
        conversation_id=conversation_id,
        playbook=body.get("playbook") or "opportunity_tree",
        artifact_mode=body.get("artifact_mode") or "report",
        ui_context={
            **(body.get("ui_context") if isinstance(body.get("ui_context"), dict) else {}),
            "entrypoint": "data_workbench",
            "selected_file_ids": [item["id"] for item in selected],
            "selected_files": selected,
            "selected_file_status": {item["id"]: item.get("status") for item in selected},
            "workbench_warnings": warnings,
        },
    )
    if background_tasks is not None:
        background_tasks.add_task(_consume_workbench_analysis, req)
        consume_result: dict[str, Any] = {"events": [], "final": None, "conversation_id": conversation_id}
    else:
        consume_result = await _consume_workbench_analysis(req, collect=True)
    resolved_conversation_id = str(consume_result.get("conversation_id") or conversation_id)
    final = consume_result.get("final") if isinstance(consume_result.get("final"), dict) else None
    return {
        "workspace_id": workspace_id,
        "conversation_id": resolved_conversation_id,
        "status": "started",
        "mode": "analysis",
        "selected_files": selected,
        "warnings": warnings,
        "jump": {"view": "agent_flow", "conversation_id": resolved_conversation_id},
        "background": True,
        "final": final,
        "text": str((final or {}).get("text") or ""),
        "events": consume_result.get("events") if isinstance(consume_result.get("events"), list) else [],
    }


async def _consume_workbench_analysis(req: ChatRequest, *, collect: bool = False) -> dict[str, Any]:
    events: list[dict[str, Any]] = []
    final: dict[str, Any] | None = None
    conversation_id = req.conversation_id
    try:
        async for chunk in orchestrate_chat(req):
            if not collect:
                continue
            for event in _parse_sse_chunk(chunk):
                events.append(event)
                data = event.get("data")
                if event.get("event") == "ready" and isinstance(data, dict) and data.get("conversation_id"):
                    conversation_id = str(data["conversation_id"])
                if event.get("event") == "final" and isinstance(data, dict):
                    final = data
    except Exception as exc:
        print(f"workbench analyze background failed: {type(exc).__name__}", flush=True)
    return {"conversation_id": conversation_id, "events": events, "final": final}


def _parse_sse_chunk(chunk: Any) -> list[dict[str, Any]]:
    frames: list[dict[str, Any]] = []
    for raw_frame in str(chunk or "").split("\n\n"):
        event_name = "message"
        data_lines: list[str] = []
        for raw_line in raw_frame.splitlines():
            line = raw_line.strip()
            if line.startswith("event:"):
                event_name = line.removeprefix("event:").strip() or event_name
            elif line.startswith("data:"):
                data_lines.append(line.removeprefix("data:").strip())
        if not data_lines:
            continue
        data_text = "\n".join(data_lines)
        try:
            data: Any = json.loads(data_text)
        except json.JSONDecodeError:
            data = data_text
        frames.append({"event": event_name, "data": data})
    return frames


def connect_sql(workspace_id: str, body: dict[str, Any]) -> dict[str, Any]:
    payload = _clean_sql_payload(body)
    connection_id = _CONNECTORS.put("sql", workspace_id, payload)
    try:
        tables = _sql_tables(payload)
    except RuntimeError as exc:
        _CONNECTORS.delete(connection_id)
        _raise_sql_connector_error(exc, status_code=503)
    except Exception as exc:
        _CONNECTORS.delete(connection_id)
        _raise_sql_connector_error(exc)
    return {
        "workspace_id": workspace_id,
        "connection_id": connection_id,
        "status": "connected",
        "expires_at": _CONNECTORS.expires_at(connection_id),
        "tables": tables,
        "credential_echo": None,
    }


def connector_status(workspace_id: str, kind: str, connection_id: str) -> dict[str, Any]:
    return _CONNECTORS.status(connection_id, kind, workspace_id)


def disconnect_connector(workspace_id: str, body: dict[str, Any]) -> dict[str, Any]:
    connection_id = str(body.get("connection_id") or "")
    kind = str(body.get("kind") or "").strip().lower()
    if kind not in {"sql", "blob"}:
        raise ValueError("kind must be sql or blob")
    if not connection_id:
        raise ValueError("connection_id is required")
    deleted = _CONNECTORS.delete(connection_id, kind=kind, workspace_id=workspace_id)
    return {"workspace_id": workspace_id, "kind": kind, "connection_id": connection_id, "disconnected": deleted}


def list_sql_tables(workspace_id: str, connection_id: str) -> dict[str, Any]:
    payload = _CONNECTORS.get(connection_id, "sql", workspace_id)
    try:
        tables = _sql_tables(payload)
    except RuntimeError as exc:
        _raise_sql_connector_error(exc, status_code=503)
    except Exception as exc:
        _raise_sql_connector_error(exc)
    return {"workspace_id": workspace_id, "connection_id": connection_id, "tables": tables}


def preview_sql_table(workspace_id: str, connection_id: str, table: str, limit: int = 100) -> dict[str, Any]:
    payload = _CONNECTORS.get(connection_id, "sql", workspace_id)
    try:
        return _sql_preview(payload, table, limit)
    except ValueError:
        raise
    except RuntimeError as exc:
        _raise_sql_connector_error(exc, status_code=503)
    except Exception as exc:
        _raise_sql_connector_error(exc)
    raise RuntimeError("unreachable")


def import_sql_table(workspace_id: str, body: dict[str, Any], actor: dict[str, Any] | None = None) -> dict[str, Any]:
    connection_id = str(body.get("connection_id") or "")
    table = str(body.get("table") or "")
    limit = _clamp_int(body.get("limit"), 1, 10000, 1000)
    preview = preview_sql_table(workspace_id, connection_id, table, limit)
    csv_bytes = _table_to_csv_bytes([col["name"] for col in preview["columns"]], preview["rows"])
    name = str(body.get("name") or f"SQL {table}").strip()[:120] or f"SQL {table}"
    result = create_workspace_upload_job(
        files=[{"filename": f"{_safe_filename(name)}.csv", "content": csv_bytes, "content_type": "text/csv"}],
        name=None,
        requested_workspace_id=workspace_id,
        asset_role="reference",
    )
    _run_ingest_if_present(result)
    _record_import_history(workspace_id, result, actor, f"Imported SQL table {table}")
    return {"workspace_id": workspace_id, "connection_id": connection_id, "imported": True, "source": {"kind": "sql", "table": table}, "upload": result}


def connect_blob(workspace_id: str, body: dict[str, Any]) -> dict[str, Any]:
    payload = _clean_blob_payload(body)
    connection_id = _CONNECTORS.put("blob", workspace_id, payload)
    try:
        containers = _blob_containers(payload)
    except Exception as exc:
        _CONNECTORS.delete(connection_id)
        raise HTTPException(status_code=400, detail=_safe_connector_error(exc)) from exc
    return {
        "workspace_id": workspace_id,
        "connection_id": connection_id,
        "status": "connected",
        "expires_at": _CONNECTORS.expires_at(connection_id),
        "containers": containers,
        "credential_echo": None,
    }


def list_blob_containers(workspace_id: str, connection_id: str) -> dict[str, Any]:
    payload = _CONNECTORS.get(connection_id, "blob", workspace_id)
    return {"workspace_id": workspace_id, "connection_id": connection_id, "containers": _blob_containers(payload)}


def list_blob_items(workspace_id: str, connection_id: str, container: str, prefix: str = "", limit: int = 100) -> dict[str, Any]:
    payload = _CONNECTORS.get(connection_id, "blob", workspace_id)
    client = _blob_service(payload).get_container_client(container)
    blobs = []
    for item in client.list_blobs(name_starts_with=prefix or ""):
        blobs.append(
            {
                "name": item.name,
                "bytes": int(getattr(item, "size", 0) or 0),
                "updated_at": _iso(getattr(item, "last_modified", None)),
                "content_type": getattr(getattr(item, "content_settings", None), "content_type", None),
            }
        )
        if len(blobs) >= limit:
            break
    return {"workspace_id": workspace_id, "connection_id": connection_id, "container": container, "prefix": prefix, "blobs": blobs}


def preview_blob_item(workspace_id: str, connection_id: str, container: str, blob: str, limit: int = 100, offset: int = 0) -> dict[str, Any]:
    payload = _CONNECTORS.get(connection_id, "blob", workspace_id)
    blob_client = _blob_service(payload).get_blob_client(container=container, blob=blob)
    props = blob_client.get_blob_properties()
    size = int(getattr(props, "size", 0) or 0)
    if size > MAX_CONNECTOR_IMPORT_BYTES:
        raise ValueError("Blob is too large for preview")
    content = blob_client.download_blob().readall()
    content_type = getattr(props.content_settings, "content_type", None)
    result = _preview_bytes(content, _type_from_name(blob, content_type), Path(blob).name, limit, offset)
    result["source"] = {"kind": "blob", "container": container, "blob": blob, "bytes": size}
    return result


def import_blob_item(workspace_id: str, body: dict[str, Any], actor: dict[str, Any] | None = None) -> dict[str, Any]:
    connection_id = str(body.get("connection_id") or "")
    container = str(body.get("container") or "")
    blob = str(body.get("blob") or "")
    if not connection_id or not container or not blob:
        raise ValueError("connection_id, container and blob are required")
    payload = _CONNECTORS.get(connection_id, "blob", workspace_id)
    blob_client = _blob_service(payload).get_blob_client(container=container, blob=blob)
    props = blob_client.get_blob_properties()
    size = int(getattr(props, "size", 0) or 0)
    if size > MAX_CONNECTOR_IMPORT_BYTES:
        raise ValueError("Blob is too large to import through the demo connector")
    content = blob_client.download_blob().readall()
    filename = Path(blob).name or "blob-import"
    content_type = getattr(props.content_settings, "content_type", None) or "application/octet-stream"
    result = create_workspace_upload_job(
        files=[{"filename": filename, "content": content, "content_type": content_type}],
        name=None,
        requested_workspace_id=workspace_id,
        asset_role="reference",
    )
    _run_ingest_if_present(result)
    _record_import_history(workspace_id, result, actor, f"Imported Blob {blob}")
    return {"workspace_id": workspace_id, "connection_id": connection_id, "imported": True, "source": {"kind": "blob", "container": container, "blob": blob}, "upload": result}


async def _call(func: Any, *args: Any, **kwargs: Any) -> Any:
    try:
        return await run_in_threadpool(func, *args, **kwargs)
    except FileNotFoundError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except PermissionError as exc:
        raise HTTPException(status_code=403, detail=str(exc)) from exc


def _require(request: Request, workspace_id: str, action: str) -> str:
    try:
        return require_workspace_permission(workspace_id, actor_from_request(request), action)
    except PermissionError as exc:
        raise HTTPException(status_code=403, detail=str(exc)) from exc


async def _json_body(request: Request) -> dict[str, Any]:
    try:
        data = await request.json()
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Invalid JSON body") from exc
    if not isinstance(data, dict):
        raise HTTPException(status_code=400, detail="JSON body must be an object")
    return data


def _file_entry(document: dict[str, Any], *, field_count: int | None = None) -> dict[str, Any]:
    file_type = _file_type(document)
    records = document.get("record_count")
    fields = document.get("field_count") or document.get("workbench_cols") or field_count
    return {
        "id": _file_id(document),
        "name": str(document.get("name") or Path(str(document.get("source_file") or "file")).name),
        "type": file_type,
        "bytes": int(document.get("bytes") or 0),
        "status": _api_status(document),
        "updated_at": document.get("updated_at") or document.get("created_at"),
        "source_file": document.get("source_file"),
        "record_count": records,
        "records": records,
        "field_count": fields,
        "fields": fields,
        "backend_status": document.get("status"),
    }


def _document_field_counts(detail: dict[str, Any], documents: list[dict[str, Any]]) -> dict[str, int]:
    columns = [item for item in detail.get("columns") or [] if isinstance(item, dict)]
    if not columns or not documents:
        return {}
    if len(documents) == 1:
        source = str(documents[0].get("source_file") or "")
        return {source: int(detail.get("field_count") or len(columns))}

    counts: dict[str, int] = {}
    for document in documents:
        source = str(document.get("source_file") or "")
        stem = Path(source).stem.lower()
        if not source or not stem:
            continue
        matched = [
            column
            for column in columns
            if stem in str(column.get("table") or "").lower()
            or str(column.get("table") or "").lower() in {stem, Path(source).name.lower()}
        ]
        if matched:
            counts[source] = len({str(column.get("name") or "") for column in matched if column.get("name")})
    return counts


def _fill_file_counts_from_content(workspace_id: str, document: dict[str, Any], entry: dict[str, Any]) -> None:
    file_type = _file_type(document)
    needs_records = entry.get("records") is None
    needs_fields = entry.get("fields") in (None, 0)
    if not needs_records and not needs_fields:
        return
    if file_type in TABLE_TYPES:
        try:
            content, _ = _read_document_bytes(workspace_id, document)
            preview = _preview_bytes(content, file_type, str(entry.get("name") or ""), limit=1, offset=0)
        except Exception:
            return
        records = preview.get("total_rows")
        fields = preview.get("total_cols")
        if needs_records and records is not None:
            entry["record_count"] = records
            entry["records"] = records
        if needs_fields and fields is not None:
            entry["field_count"] = fields
            entry["fields"] = fields
    elif file_type in MARKDOWN_TYPES:
        if needs_records:
            entry["record_count"] = 1
            entry["records"] = 1
        if needs_fields:
            entry["field_count"] = 1
            entry["fields"] = 1
    elif file_type in JSON_TYPES:
        try:
            content, _ = _read_document_bytes(workspace_id, document)
            table = _json_table(content)
        except Exception:
            return
        if needs_records:
            entry["record_count"] = len(table.get("rows") or [])
            entry["records"] = entry["record_count"]
        if needs_fields:
            entry["field_count"] = len(table.get("headers") or [])
            entry["fields"] = entry["field_count"]


def _file_id(document: dict[str, Any]) -> str:
    source = str(document.get("source_file") or document.get("name") or "file")
    return hashlib.sha256(source.encode("utf-8")).hexdigest()[:16]


def _file_type(document: dict[str, Any]) -> str:
    name = str(document.get("name") or document.get("source_file") or "")
    suffix = Path(name).suffix.lower().lstrip(".")
    fmt = str(document.get("format") or "").lower()
    if suffix:
        return "xlsx" if suffix == "xlsm" else suffix
    if fmt == "excel":
        return "xlsx"
    if fmt == "markdown":
        return "md"
    return fmt or "unknown"


def _file_group(entry: dict[str, Any]) -> str:
    name = str(entry.get("name") or "").lower()
    file_type = str(entry.get("type") or "").lower()
    if name.startswith(("temp", "tmp", "untitled")):
        return "临时文件"
    if file_type in {"csv", "xlsx", "xlsm", "json"}:
        return "数据集"
    return "文档"


def _api_status(document: dict[str, Any]) -> str:
    status = str(document.get("status") or "").lower()
    if "ready" in status or "indexed" in status or "就绪" in status:
        return "indexed"
    if _file_type(document) in MARKDOWN_TYPES | JSON_TYPES and not document.get("error"):
        return "indexed"
    return "needs_review"


def _find_document(workspace_id: str, file_id: str, *, detail: dict[str, Any] | None = None) -> dict[str, Any]:
    detail = detail or get_workspace_detail(workspace_id)
    target = str(file_id or "").strip()
    for item in detail.get("documents") or []:
        if not isinstance(item, dict):
            continue
        if target in {_file_id(item), str(item.get("source_file") or ""), str(item.get("name") or "")}:
            return dict(item)
    raise FileNotFoundError(f"File not found: {file_id}")


def _workspace_state(workspace_id: str) -> tuple[dict[str, Any], dict[str, Any]]:
    loaded = _load_workspace_bundle(workspace_id)
    if not loaded:
        raise FileNotFoundError(workspace_id)
    return loaded


def _read_document_bytes(workspace_id: str, document: dict[str, Any]) -> tuple[bytes, str]:
    source = _safe_source_file(document)
    local_path = WORKSPACES / workspace_id / source
    if local_path.exists() and local_path.is_file():
        return local_path.read_bytes(), str(document.get("content_type") or "application/octet-stream")
    blob = download_blob_content(f"workspaces/{workspace_id}/{source}")
    if blob:
        return blob
    raise FileNotFoundError(f"File content not found: {document.get('name') or source}")


def _write_document_bytes(workspace_id: str, document: dict[str, Any], content: bytes, content_type: str) -> None:
    source = _safe_source_file(document)
    local_path = WORKSPACES / workspace_id / source
    local_path.parent.mkdir(parents=True, exist_ok=True)
    local_path.write_bytes(content)
    try:
        upload_workspace_blob(workspace_id, source, content, content_type)
    except Exception:
        pass


def _safe_source_file(document: dict[str, Any]) -> str:
    source = str(document.get("source_file") or "").replace("\\", "/")
    if not source or source.startswith("/") or ".." in Path(source).parts or not source.startswith("raw_docs/"):
        raise ValueError("Only workspace raw_docs files can be edited or previewed")
    return source


def _preview_bytes(content: bytes, file_type: str, name: str, limit: int, offset: int) -> dict[str, Any]:
    if file_type in {"csv"}:
        return _preview_csv(content, limit, offset)
    if file_type in {"xlsx", "xlsm", "excel"}:
        return _preview_xlsx(content, limit, offset)
    if file_type in MARKDOWN_TYPES:
        text = _decode_text(content)
        end = min(len(text), offset + min(limit * 1200, MAX_MARKDOWN_BYTES))
        return {"kind": "markdown", "text": text[offset:end], "total_chars": len(text), "offset": offset, "limit": limit}
    if file_type in JSON_TYPES:
        return _preview_json(content, limit, offset)
    return {"kind": "binary", "name": name, "bytes": len(content), "message": "Preview is not available for this file type"}


def _preview_csv(content: bytes, limit: int, offset: int) -> dict[str, Any]:
    rows = _csv_rows(content)
    if not rows:
        return {"kind": "table", "columns": [], "rows": [], "total_rows": 0, "total_cols": 0, "offset": offset, "limit": limit}
    width = max(len(row) for row in rows)
    headers = _headers(rows[0], width)
    body = [_pad(row, width) for row in rows[1:]]
    return {
        "kind": "table",
        "columns": [{"name": name} for name in headers],
        "rows": body[offset : offset + limit],
        "total_rows": len(body),
        "total_cols": width,
        "offset": offset,
        "limit": limit,
    }


def _preview_xlsx(content: bytes, limit: int, offset: int) -> dict[str, Any]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        sheet = workbook.worksheets[0]
        rows_iter = sheet.iter_rows(values_only=True)
        try:
            first = list(next(rows_iter))
        except StopIteration:
            return {"kind": "table", "columns": [], "rows": [], "total_rows": 0, "total_cols": 0, "offset": offset, "limit": limit, "sheet": sheet.title}
        width = max(sheet.max_column or len(first), len(first))
        headers = _headers(first, width)
        rows: list[list[str]] = []
        for idx, row in enumerate(rows_iter):
            if idx < offset:
                continue
            if len(rows) >= limit:
                break
            rows.append(_pad([_cell(value) for value in row], width))
        total_rows = max(0, int(sheet.max_row or 1) - 1)
        return {
            "kind": "table",
            "columns": [{"name": name} for name in headers],
            "rows": rows,
            "total_rows": total_rows,
            "total_cols": width,
            "offset": offset,
            "limit": limit,
            "sheet": sheet.title,
            "sheets": [ws.title for ws in workbook.worksheets],
        }
    finally:
        workbook.close()


def _preview_json(content: bytes, limit: int, offset: int) -> dict[str, Any]:
    raw = _decode_text(content)
    try:
        value = json.loads(raw)
        text = json.dumps(value, ensure_ascii=False, indent=2)
        table = _json_table(content)
        return {
            "kind": "json",
            "text": text[offset : min(len(text), offset + min(limit * 1200, MAX_MARKDOWN_BYTES))],
            "total_chars": len(text),
            "offset": offset,
            "limit": limit,
            "shape": {"rows": len(table.get("rows") or []), "fields": len(table.get("headers") or [])},
        }
    except json.JSONDecodeError as exc:
        return {
            "kind": "json",
            "text": raw[offset : min(len(raw), offset + min(limit * 1200, MAX_MARKDOWN_BYTES))],
            "total_chars": len(raw),
            "offset": offset,
            "limit": limit,
            "parse_error": f"{exc.msg} at line {exc.lineno}, column {exc.colno}",
        }


def _table_from_bytes(content: bytes, file_type: str) -> dict[str, Any]:
    if file_type == "csv":
        preview = _preview_csv(content, limit=10_000_000, offset=0)
        return {"headers": [col["name"] for col in preview["columns"]], "rows": preview["rows"]}
    if file_type in {"xlsx", "xlsm", "excel"}:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        try:
            sheet = workbook.worksheets[0]
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                return {"headers": [], "rows": []}
            width = max(len(row) for row in rows)
            return {"headers": _headers(list(rows[0]), width), "rows": [_pad([_cell(value) for value in row], width) for row in rows[1:]]}
        finally:
            workbook.close()
    return {"headers": [], "rows": []}


def _json_table(content: bytes) -> dict[str, Any]:
    value = json.loads(_decode_text(content))
    records: list[dict[str, Any]]
    if isinstance(value, list):
        if all(isinstance(item, dict) for item in value):
            records = [item for item in value if isinstance(item, dict)]
        else:
            records = [{"value": item} for item in value]
    elif isinstance(value, dict):
        list_key = next(
            (
                key
                for key, item in value.items()
                if isinstance(item, list) and item and all(isinstance(row, dict) for row in item)
            ),
            None,
        )
        if list_key:
            records = [item for item in value[list_key] if isinstance(item, dict)]
        else:
            records = [value]
    else:
        records = [{"value": value}]

    headers: list[str] = []
    seen: set[str] = set()
    for record in records:
        for key in record.keys():
            name = str(key or "").strip() or "value"
            if name not in seen:
                seen.add(name)
                headers.append(name)
    if not headers:
        headers = ["value"]
    return {"headers": headers, "rows": [[_json_cell(record.get(header)) for header in headers] for record in records]}


def _json_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    return _cell(value)


def _new_file_name_and_type(body: dict[str, Any]) -> tuple[str, str]:
    raw_name = str(body.get("name") or body.get("filename") or "").strip()
    requested = str(body.get("type") or body.get("kind") or "").strip().lower()
    if requested in {"markdown", "text", "note"}:
        requested = "md"
    elif requested in {"table", "spreadsheet"}:
        requested = "csv"
    elif requested == "excel":
        requested = "xlsx"

    if not raw_name:
        raw_name = "untitled"
    if "/" in raw_name or "\\" in raw_name or raw_name in {".", ".."}:
        raise ValueError("Invalid file name")
    if len(raw_name) > 160:
        raise ValueError("File name is too long")

    name_path = Path(raw_name)
    suffix = name_path.suffix.lower().lstrip(".")
    if suffix == "markdown":
        suffix = "md"
    if suffix == "excel":
        suffix = "xlsx"
    file_type = requested or suffix or "md"
    if file_type not in {"md", "txt", "csv", "xlsx"}:
        raise ValueError("type must be one of md, txt, csv or xlsx")
    if file_type == "txt":
        ext = ".txt"
    elif file_type == "md":
        ext = ".md"
    elif file_type == "xlsx":
        ext = ".xlsx"
    else:
        ext = ".csv"

    stem = name_path.stem if name_path.suffix else raw_name
    stem = re.sub(r"[\x00-\x1f<>:\"|?*]+", "-", stem).strip(" .-_")
    if not stem:
        stem = "untitled"
    return f"{stem[:120]}{ext}", file_type


def _new_file_content(body: dict[str, Any], file_type: str) -> tuple[bytes, str]:
    if file_type in {"md", "txt"}:
        text = body.get("text")
        if text is None:
            text = body.get("content")
        if text is None:
            text = "# Untitled\n"
        if not isinstance(text, str):
            raise ValueError("text/content must be a string")
        content = text.encode("utf-8")
        if len(content) > MAX_MARKDOWN_BYTES:
            raise ValueError("Markdown content is too large")
        content_type = "text/plain; charset=utf-8" if file_type == "txt" else "text/markdown; charset=utf-8"
        return content or b"\n", content_type

    columns, rows = _new_table_shape(body)
    if file_type == "csv":
        content = _table_to_csv_bytes(columns, rows)
        if len(content) > MAX_TABLE_SAVE_BYTES:
            raise ValueError("Table content is too large")
        return content, "text/csv; charset=utf-8"

    workbook = Workbook()
    try:
        sheet = workbook.active
        sheet.title = "Sheet1"
        sheet.append(columns)
        for row in rows:
            sheet.append(row)
        out = io.BytesIO()
        workbook.save(out)
        content = out.getvalue()
    finally:
        workbook.close()
    if len(content) > MAX_TABLE_SAVE_BYTES:
        raise ValueError("Table content is too large")
    return content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _new_table_shape(body: dict[str, Any]) -> tuple[list[str], list[list[str]]]:
    if "columns" in body:
        raw_columns = body.get("columns")
    elif "headers" in body:
        raw_columns = body.get("headers")
    else:
        raw_columns = ["column_1"]
    if not isinstance(raw_columns, list) or not raw_columns:
        raise ValueError("columns must be a non-empty list")
    if len(raw_columns) > 200:
        raise ValueError("Too many columns")
    columns = [_column_name(item, idx) for idx, item in enumerate(raw_columns)]

    raw_rows = body.get("rows") if "rows" in body else []
    if not isinstance(raw_rows, list):
        raise ValueError("rows must be a list")
    if len(raw_rows) > 5000:
        raise ValueError("Too many rows")
    rows: list[list[str]] = []
    for raw_row in raw_rows:
        if isinstance(raw_row, dict):
            row = [str(raw_row.get(column, "")) for column in columns]
        elif isinstance(raw_row, list):
            row = [str(raw_row[idx]) if idx < len(raw_row) and raw_row[idx] is not None else "" for idx in range(len(columns))]
        else:
            raise ValueError("Each row must be an array or object")
        for value in row:
            if len(value) > MAX_CELL_CHARS:
                raise ValueError("Cell value is too long")
        rows.append(row)
    return columns, rows


def _column_name(value: Any, index: int) -> str:
    name = str(value or "").strip() or f"column_{index + 1}"
    name = re.sub(r"\s+", " ", name)
    if len(name) > 120:
        raise ValueError("Column name is too long")
    return name


def _compact_upload_result(result: dict[str, Any]) -> dict[str, Any]:
    documents = []
    for item in result.get("documents") or []:
        if not isinstance(item, dict):
            continue
        documents.append(
            {
                "id": _file_id(item),
                "name": item.get("name"),
                "type": _file_type(item),
                "source_file": item.get("source_file"),
                "bytes": item.get("bytes"),
                "status": _api_status(item),
                "updated_at": item.get("updated_at") or item.get("created_at"),
            }
        )
    return {
        "workspace_id": result.get("workspace_id"),
        "ingest_job_id": result.get("ingest_job_id"),
        "ingest_status": result.get("ingest_status"),
        "indexed_count": result.get("indexed_count"),
        "indexed_delta": result.get("indexed_delta"),
        "documents": documents,
    }


def _quality_from_json(workspace_id: str, document: dict[str, Any], content: bytes, checked_at: str) -> dict[str, Any]:
    try:
        table = _json_table(content)
    except json.JSONDecodeError as exc:
        return {
            "workspace_id": workspace_id,
            "file_id": _file_id(document),
            "field_mapping": {"mapped": 0, "total": 0, "pct": 0.0, "fields": []},
            "quality": {
                "missing_pct": 100.0,
                "duplicate_pct": 0.0,
                "outlier_count": 0,
                "outlier_pct": 0.0,
                "type_warnings": 1,
                "row_count": 0,
                "column_count": 0,
                "parse_error": f"{exc.msg} at line {exc.lineno}, column {exc.colno}",
            },
            "validation": {"status": "failed", "checked_at": checked_at},
        }
    return _quality_from_table(workspace_id, document, table, checked_at)


def _quality_from_table(workspace_id: str, document: dict[str, Any], table: dict[str, Any], checked_at: str) -> dict[str, Any]:
    headers = [str(item) for item in table.get("headers") or []]
    rows = [[str(cell) for cell in row] for row in (table.get("rows") or [])]
    total_cols = len(headers)
    total_rows = len(rows)
    total_cells = max(1, total_rows * max(1, total_cols))
    missing = sum(1 for row in rows for value in _pad(row, total_cols) if not str(value).strip())
    duplicate_count = _duplicate_count(rows)
    fields = [_field_quality(idx, name, rows) for idx, name in enumerate(headers)]
    mapped = sum(1 for item in fields if item["mapped"])
    outlier_count = sum(int(item.get("outlier_count") or 0) for item in fields)
    type_warnings = sum(1 for item in fields if item.get("type_warning"))
    missing_pct = round(missing / total_cells * 100, 2)
    duplicate_pct = round(duplicate_count / max(1, total_rows) * 100, 2)
    outlier_pct = round(outlier_count / max(1, total_rows * max(1, total_cols)) * 100, 2)
    if total_rows == 0 or total_cols == 0 or missing_pct >= 50:
        status = "failed"
    elif missing_pct > 5 or duplicate_pct > 5 or outlier_count > 0 or type_warnings > 0:
        status = "warn"
    else:
        status = "passed"
    return {
        "workspace_id": workspace_id,
        "file_id": _file_id(document),
        "field_mapping": {"mapped": mapped, "total": total_cols, "pct": round(mapped / total_cols * 100, 2) if total_cols else 0.0, "fields": fields},
        "quality": {
            "missing_pct": missing_pct,
            "duplicate_pct": duplicate_pct,
            "outlier_count": outlier_count,
            "outlier_pct": outlier_pct,
            "type_warnings": type_warnings,
            "row_count": total_rows,
            "column_count": total_cols,
        },
        "validation": {"status": status, "checked_at": checked_at},
    }


def _apply_mapping_overrides(workspace_id: str, file_id: str, result: dict[str, Any]) -> dict[str, Any]:
    meta, _profile = _workspace_state(workspace_id)
    stored = _stored_field_mapping(meta, file_id)
    mappings = stored.get("mappings") or {}
    field_mapping = result.get("field_mapping") if isinstance(result.get("field_mapping"), dict) else {}
    fields = field_mapping.get("fields") if isinstance(field_mapping.get("fields"), list) else []
    mapped = 0
    for field in fields:
        if not isinstance(field, dict):
            continue
        name = str(field.get("name") or "")
        override = mappings.get(name) if isinstance(mappings, dict) else None
        target = override.get("target") if isinstance(override, dict) else None
        if target:
            field["mapped"] = True
            field["target"] = target
            field["standard_name"] = target
            field["mapping_source"] = "user"
            if override.get("notes"):
                field["notes"] = override.get("notes")
        if field.get("mapped"):
            mapped += 1
    total = int(field_mapping.get("total") or len([item for item in fields if isinstance(item, dict)]))
    field_mapping["mapped"] = mapped
    field_mapping["total"] = total
    field_mapping["pct"] = round(mapped / total * 100, 2) if total else 0.0
    if stored.get("saved_at"):
        field_mapping["overrides_saved_at"] = stored.get("saved_at")
    result["field_mapping"] = field_mapping
    return result


def _stored_field_mapping(meta: dict[str, Any], file_id: str) -> dict[str, Any]:
    all_mappings = meta.get("workbench_field_mappings")
    if not isinstance(all_mappings, dict):
        return {}
    stored = all_mappings.get(str(file_id))
    return stored if isinstance(stored, dict) else {}


def _normalize_field_mapping_body(body: dict[str, Any], fields: set[str]) -> dict[str, dict[str, str]]:
    raw = body.get("mappings", body.get("mapping", body.get("overrides", {})))
    pairs: list[tuple[str, Any, Any]] = []
    if isinstance(raw, dict):
        for source, target in raw.items():
            notes = None
            if isinstance(target, dict):
                notes = target.get("notes")
                target = target.get("target", target.get("standard_name", target.get("mapped_to")))
            pairs.append((str(source), target, notes))
    elif isinstance(raw, list):
        for item in raw:
            if not isinstance(item, dict):
                raise ValueError("Each mapping must be an object")
            source = str(item.get("name") or item.get("source") or item.get("field") or "").strip()
            target = item.get("target", item.get("standard_name", item.get("mapped_to")))
            notes = item.get("notes")
            pairs.append((source, target, notes))
    else:
        raise ValueError("mappings must be an object or list")

    normalized: dict[str, dict[str, str]] = {}
    for source, target, notes in pairs:
        source = str(source or "").strip()
        if not source:
            raise ValueError("Mapping source field is required")
        if fields and source not in fields:
            raise ValueError(f"Unknown field: {source}")
        target_text = str(target or "").strip()
        if not target_text:
            continue
        if len(target_text) > 120:
            raise ValueError("Mapping target is too long")
        entry = {"target": target_text}
        notes_text = str(notes or "").strip()
        if notes_text:
            if len(notes_text) > 300:
                raise ValueError("Mapping notes are too long")
            entry["notes"] = notes_text
        normalized[source] = entry
    return normalized


def _field_quality(idx: int, name: str, rows: list[list[str]]) -> dict[str, Any]:
    values = [str(row[idx]).strip() if idx < len(row) else "" for row in rows]
    present = [value for value in values if value]
    inferred, numbers = _infer_type(present)
    non_empty = len(present)
    mixed_ratio = _mixed_ratio(present, inferred)
    outliers = _outliers(numbers)
    return {
        "name": name,
        "type": inferred,
        "mapped": bool(name.strip()) and non_empty > 0 and inferred != "empty",
        "missing_pct": round((len(values) - non_empty) / max(1, len(values)) * 100, 2),
        "unique_count": len(set(present)),
        "type_warning": mixed_ratio > 15,
        "mixed_type_pct": round(mixed_ratio, 2),
        "outlier_count": len(outliers),
    }


def _table_operations_from_body(body: dict[str, Any]) -> dict[str, Any]:
    aliases = {
        "add_rows": ("add_rows", "insert_rows"),
        "delete_rows": ("delete_rows", "remove_rows"),
        "add_cols": ("add_cols", "insert_cols", "add_columns", "insert_columns"),
        "delete_cols": ("delete_cols", "remove_cols", "delete_columns", "remove_columns"),
    }
    operations: dict[str, Any] = {}
    for target, keys in aliases.items():
        for key in keys:
            if key in body and body.get(key) not in (None, "", [], {}):
                operations[target] = body.get(key)
                break
    return operations


def _apply_csv_table_update(content: bytes, edits: list[Any], operations: dict[str, Any]) -> tuple[bytes, int, int, dict[str, int]]:
    rows = _csv_rows(content)
    if not rows:
        raise ValueError("CSV file has no rows")
    width = max(len(row) for row in rows)
    headers = _headers(rows[0], width)
    data_rows = [_pad(row, width) for row in rows[1:]]
    headers, data_rows, counts = _apply_table_update(headers, data_rows, edits, operations)
    return _table_to_csv_bytes(headers, data_rows), len(data_rows), len(headers), counts


def _apply_xlsx_table_update(content: bytes, edits: list[Any], operations: dict[str, Any]) -> tuple[bytes, int, int, dict[str, int]]:
    table = _table_from_bytes(content, "xlsx")
    headers, data_rows, counts = _apply_table_update(table.get("headers") or [], table.get("rows") or [], edits, operations)
    workbook = Workbook()
    try:
        sheet = workbook.active
        sheet.title = "Sheet1"
        sheet.append(headers)
        for row in data_rows:
            sheet.append(row)
        out = io.BytesIO()
        workbook.save(out)
        return out.getvalue(), len(data_rows), len(headers), counts
    finally:
        workbook.close()


def _apply_table_update(
    headers: list[str],
    rows: list[list[Any]],
    edits: list[Any],
    operations: dict[str, Any],
) -> tuple[list[str], list[list[str]], dict[str, int]]:
    headers = [_column_name(name, idx) for idx, name in enumerate(headers or ["column_1"])]
    rows = [_pad(list(row), len(headers)) for row in rows]
    counts = {"cells": len(edits), "rows_added": 0, "rows_deleted": 0, "cols_added": 0, "cols_deleted": 0}

    for row_idx in sorted(_row_indexes(operations.get("delete_rows"), len(rows)), reverse=True):
        del rows[row_idx]
        counts["rows_deleted"] += 1

    for col_idx in sorted(_col_indexes(operations.get("delete_cols"), headers), reverse=True):
        del headers[col_idx]
        for row in rows:
            del row[col_idx]
        counts["cols_deleted"] += 1
    if not headers:
        headers = ["column_1"]
        for row in rows:
            row.append("")

    for item in _row_inserts(operations.get("add_rows"), headers):
        idx, values = item
        idx = max(0, min(len(rows), idx))
        rows.insert(idx, values)
        counts["rows_added"] += 1

    for item in _col_inserts(operations.get("add_cols"), headers, len(rows)):
        idx, name, values = item
        idx = max(0, min(len(headers), idx))
        unique = _unique_column_name(name, headers)
        headers.insert(idx, unique)
        for row_idx, row in enumerate(rows):
            row.insert(idx, values[row_idx] if row_idx < len(values) else "")
        counts["cols_added"] += 1

    if len(edits) > MAX_CELL_EDITS:
        raise ValueError(f"Too many edits; max {MAX_CELL_EDITS}")
    normalized = [_normalize_edit(edit, headers, len(rows), len(headers)) for edit in edits]
    for row_idx, col_idx, value in normalized:
        rows[row_idx][col_idx] = value
    return headers, rows, counts


def _row_indexes(raw: Any, row_count: int) -> list[int]:
    if raw in (None, "", [], {}):
        return []
    values = raw if isinstance(raw, list) else [raw]
    indexes = sorted({_clamp_int(item, 0, max(0, row_count - 1), None) for item in values})
    return indexes if row_count else []


def _col_indexes(raw: Any, headers: list[str]) -> list[int]:
    if raw in (None, "", [], {}):
        return []
    values = raw if isinstance(raw, list) else [raw]
    indexes: set[int] = set()
    for item in values:
        if isinstance(item, str) and not item.isdigit():
            if item not in headers:
                raise ValueError(f"Unknown column: {item}")
            indexes.add(headers.index(item))
        else:
            indexes.add(_clamp_int(item, 0, max(0, len(headers) - 1), None))
    return sorted(indexes)


def _row_inserts(raw: Any, headers: list[str]) -> list[tuple[int, list[str]]]:
    if raw in (None, "", [], {}):
        return []
    if isinstance(raw, int):
        return [(10**9, ["" for _ in headers]) for _ in range(max(0, raw))]
    items = raw if isinstance(raw, list) else [raw]
    inserts: list[tuple[int, list[str]]] = []
    for item in items:
        idx = 10**9
        values: Any = item
        if isinstance(item, dict) and ("values" in item or "index" in item):
            idx = _clamp_int(item.get("index"), 0, 10**9, 10**9)
            values = item.get("values", {})
        inserts.append((idx, _row_values(values, headers)))
    return inserts


def _col_inserts(raw: Any, headers: list[str], row_count: int) -> list[tuple[int, str, list[str]]]:
    if raw in (None, "", [], {}):
        return []
    if isinstance(raw, int):
        return [(10**9, f"column_{len(headers) + idx + 1}", ["" for _ in range(row_count)]) for idx in range(max(0, raw))]
    items = raw if isinstance(raw, list) else [raw]
    inserts: list[tuple[int, str, list[str]]] = []
    for idx, item in enumerate(items):
        if isinstance(item, dict):
            name = _column_name(item.get("name") or item.get("header"), len(headers) + idx)
            insert_at = _clamp_int(item.get("index"), 0, 10**9, 10**9)
            values = item.get("values") if isinstance(item.get("values"), list) else []
            inserts.append((insert_at, name, [str(value if value is not None else "")[:MAX_CELL_CHARS] for value in values]))
        else:
            inserts.append((10**9, _column_name(item, len(headers) + idx), ["" for _ in range(row_count)]))
    return inserts


def _row_values(raw: Any, headers: list[str]) -> list[str]:
    if isinstance(raw, dict):
        values = [raw.get(header, "") for header in headers]
    elif isinstance(raw, list):
        values = [raw[idx] if idx < len(raw) else "" for idx in range(len(headers))]
    else:
        values = ["" for _ in headers]
    result = [str(value if value is not None else "") for value in values]
    if any(len(value) > MAX_CELL_CHARS for value in result):
        raise ValueError("Cell value is too long")
    return result


def _unique_column_name(name: str, headers: list[str]) -> str:
    base = _column_name(name, len(headers))
    if base not in headers:
        return base
    idx = 2
    while f"{base}_{idx}" in headers:
        idx += 1
    return f"{base}_{idx}"


def _table_change_summary(counts: dict[str, int]) -> str:
    parts: list[str] = []
    if counts.get("cells"):
        parts.append(f"更新了 {counts['cells']} 个单元格")
    if counts.get("rows_added"):
        parts.append(f"新增 {counts['rows_added']} 行")
    if counts.get("rows_deleted"):
        parts.append(f"删除 {counts['rows_deleted']} 行")
    if counts.get("cols_added"):
        parts.append(f"新增 {counts['cols_added']} 列")
    if counts.get("cols_deleted"):
        parts.append(f"删除 {counts['cols_deleted']} 列")
    return "，".join(parts) if parts else "更新了表格结构"


def _apply_csv_edits(content: bytes, edits: list[Any]) -> tuple[bytes, int, int]:
    rows = _csv_rows(content)
    if not rows:
        raise ValueError("CSV file has no rows")
    width = max(len(row) for row in rows)
    headers = _headers(rows[0], width)
    normalized = [_normalize_edit(edit, headers, len(rows) - 1, width) for edit in edits]
    rows = [_pad(row, width) for row in rows]
    for row_idx, col_idx, value in normalized:
        rows[row_idx + 1][col_idx] = value
    out = io.StringIO()
    writer = csv.writer(out, lineterminator="\n")
    writer.writerows(rows)
    return out.getvalue().encode("utf-8"), len(rows) - 1, width


def _apply_xlsx_edits(content: bytes, edits: list[Any]) -> tuple[bytes, int, int]:
    workbook = load_workbook(io.BytesIO(content))
    try:
        sheet = workbook.worksheets[0]
        first = [sheet.cell(row=1, column=idx).value for idx in range(1, (sheet.max_column or 0) + 1)]
        headers = _headers(first, sheet.max_column or len(first))
        normalized = [_normalize_edit(edit, headers, max(0, (sheet.max_row or 1) - 1), sheet.max_column or len(headers)) for edit in edits]
        for row_idx, col_idx, value in normalized:
            sheet.cell(row=row_idx + 2, column=col_idx + 1).value = value
        out = io.BytesIO()
        workbook.save(out)
        return out.getvalue(), max(0, (sheet.max_row or 1) - 1), sheet.max_column or len(headers)
    finally:
        workbook.close()


def _normalize_edit(edit: Any, headers: list[str], row_count: int, col_count: int) -> tuple[int, int, str]:
    if not isinstance(edit, dict):
        raise ValueError("Each edit must be an object")
    row = _clamp_int(edit.get("row"), 0, max(0, row_count - 1), None)
    raw_col = edit.get("col", edit.get("column"))
    if isinstance(raw_col, str) and not raw_col.isdigit():
        if raw_col not in headers:
            raise ValueError(f"Unknown column: {raw_col}")
        col = headers.index(raw_col)
    else:
        col = _clamp_int(raw_col, 0, max(0, col_count - 1), None)
    value = str(edit.get("value") if edit.get("value") is not None else "")
    if len(value) > MAX_CELL_CHARS:
        raise ValueError("Cell value is too long")
    return row, col, value


def _snapshot_version(workspace_id: str, document: dict[str, Any], content: bytes) -> dict[str, Any]:
    version_id = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%S") + "-" + uuid.uuid4().hex[:8]
    source = _safe_source_file(document)
    filename = Path(source).name
    rel = f"versions/{_file_id(document)}/{version_id}-{filename}"
    path = WORKSPACES / workspace_id / rel
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(content)
    try:
        upload_workspace_blob(workspace_id, rel, content, str(document.get("content_type") or "application/octet-stream"))
    except Exception:
        pass
    return {"version_id": version_id, "source_file": rel, "bytes": len(content)}


def _record_import_history(workspace_id: str, result: dict[str, Any], actor: dict[str, Any] | None, summary: str) -> None:
    created_sources = {
        str(item.get("source_file") or "")
        for item in result.get("documents") or []
        if isinstance(item, dict) and item.get("ingest_job_id") == result.get("ingest_job_id")
    }
    if not created_sources:
        return
    try:
        detail = get_workspace_detail(workspace_id)
    except Exception:
        return
    for document in detail.get("documents") or []:
        if isinstance(document, dict) and str(document.get("source_file") or "") in created_sources:
            _append_workbench_history(workspace_id, _file_id(document), actor, _utc_now(), summary)


def _append_workbench_history(
    workspace_id: str,
    file_id: str,
    actor: dict[str, Any] | None,
    at: str,
    summary: str,
    *,
    version_id: str | None = None,
    previous_source_file: str | None = None,
) -> None:
    fid = str(file_id or "").strip()
    if not fid:
        return
    try:
        meta, profile = _workspace_state(workspace_id)
        history = meta.setdefault("workbench_history", {})
        items = history.setdefault(fid, [])
        item = {
            **actor_for_history(actor),
            "at": at,
            "change_summary": summary,
        }
        if version_id:
            item["version_id"] = version_id
        if previous_source_file:
            item["previous_source_file"] = previous_source_file
        items.insert(0, item)
        del items[20:]
        _persist_workspace_state(workspace_id, WORKSPACES / workspace_id, meta, profile, include_raw_payloads=True)
    except Exception:
        return


def _update_document_after_save(
    workspace_id: str,
    document: dict[str, Any],
    content: bytes,
    saved_at: str,
    summary: str,
    version: dict[str, Any],
    *,
    row_count: int,
    col_count: int,
    actor: dict[str, Any] | None = None,
) -> None:
    meta, profile = _workspace_state(workspace_id)
    fid = _file_id(document)
    source = _safe_source_file(document)
    if not isinstance(meta.get("documents"), list) or not meta.get("documents"):
        meta["documents"] = _normalize_documents(WORKSPACES / workspace_id, meta, profile)
    changed = False
    for item in meta.get("documents") or []:
        if isinstance(item, dict) and str(item.get("source_file") or "") == source:
            item.update(
                {
                    "bytes": len(content),
                    "record_count": row_count,
                    "updated_at": saved_at,
                    "content_sha256": hashlib.sha256(content).hexdigest(),
                    "status": "needs_review",
                    "workbench_cols": col_count,
                }
            )
            changed = True
            profile_file = str(item.get("profile_file") or "")
            if profile_file:
                _refresh_document_profile(workspace_id, meta, profile, item, profile_file)
            break
    if not changed:
        raise FileNotFoundError(source)
    history = meta.setdefault("workbench_history", {})
    items = history.setdefault(fid, [])
    items.insert(
        0,
        {
            **actor_for_history(actor),
            "at": saved_at,
            "change_summary": summary,
            "version_id": version["version_id"],
            "previous_source_file": version["source_file"],
        },
    )
    del items[20:]
    profile = _rebuild_workspace_profile(workspace_id, WORKSPACES / workspace_id, meta, profile)
    _persist_workspace_state(workspace_id, WORKSPACES / workspace_id, meta, profile, include_raw_payloads=True)
    _CONTEXT_CACHE.pop(workspace_id, None)


def _refresh_document_profile(workspace_id: str, meta: dict[str, Any], profile: dict[str, Any], document: dict[str, Any], profile_file: str) -> None:
    raw_path = WORKSPACES / workspace_id / _safe_source_file(document)
    item_profile = build_data_profile(
        raw_path,
        workspace_id=workspace_id,
        name=str(meta.get("name") or profile.get("name") or workspace_id),
        source_file=str(document.get("source_file") or ""),
        content_type=document.get("content_type"),
    )
    item_profile["profile_id"] = Path(profile_file).stem
    item_profile["profile_file"] = profile_file
    profile_path = WORKSPACES / workspace_id / profile_file
    profile_path.parent.mkdir(parents=True, exist_ok=True)
    profile_path.write_text(json.dumps(item_profile, ensure_ascii=False, indent=2), encoding="utf-8")
    try:
        docs = [profile_to_search_document(item_profile)] + upload_to_records(
            raw_path,
            str(document.get("source_file") or ""),
            workspace_id,
            content_type=document.get("content_type"),
        )
        document["indexed_count"] = _index_documents_batched(docs)
    except Exception as exc:
        document["index_warning"] = f"{type(exc).__name__}: {exc}"[:300]


def _csv_rows(content: bytes) -> list[list[str]]:
    text = _decode_text(content)
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample)
    except csv.Error:
        dialect = csv.excel
    return [[_cell(value) for value in row] for row in csv.reader(text.splitlines(), dialect)]


def _headers(values: list[Any], width: int) -> list[str]:
    names: list[str] = []
    seen: dict[str, int] = {}
    for idx in range(width):
        raw = _cell(values[idx] if idx < len(values) else "") or f"column_{idx + 1}"
        raw = re.sub(r"\s+", " ", raw)
        if raw in seen:
            seen[raw] += 1
            raw = f"{raw}_{seen[raw]}"
        else:
            seen[raw] = 1
        names.append(raw)
    return names


def _pad(row: list[Any], width: int) -> list[str]:
    return [_cell(row[idx] if idx < len(row) else "") for idx in range(width)]


def _cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    return str(value).strip()


def _decode_text(content: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8", errors="replace")


def _duplicate_count(rows: list[list[str]]) -> int:
    counter = Counter(tuple(row) for row in rows if any(str(cell).strip() for cell in row))
    return sum(count - 1 for count in counter.values() if count > 1)


def _infer_type(values: list[str]) -> tuple[str, list[float]]:
    if not values:
        return "empty", []
    nums = [_to_float(value) for value in values]
    numeric = [value for value in nums if value is not None]
    if len(numeric) / len(values) >= 0.85:
        return "number", numeric
    if sum(1 for value in values if value.lower() in {"true", "false", "yes", "no", "0", "1"}) / len(values) >= 0.85:
        return "boolean", []
    if sum(1 for value in values if re.match(r"^\d{4}[-/]\d{1,2}[-/]\d{1,2}", value)) / len(values) >= 0.75:
        return "date", []
    avg = sum(len(value) for value in values) / len(values)
    return ("long_text" if avg > 80 else "text"), []


def _mixed_ratio(values: list[str], inferred: str) -> float:
    if not values or inferred in {"empty", "text", "long_text"}:
        return 0.0
    mismatches = 0
    for value in values:
        if inferred == "number" and _to_float(value) is None:
            mismatches += 1
        elif inferred == "boolean" and value.lower() not in {"true", "false", "yes", "no", "0", "1"}:
            mismatches += 1
        elif inferred == "date" and not re.match(r"^\d{4}[-/]\d{1,2}[-/]\d{1,2}", value):
            mismatches += 1
    return mismatches / len(values) * 100


def _outliers(numbers: list[float]) -> list[float]:
    if len(numbers) < 8:
        return []
    ordered = sorted(numbers)
    q1 = ordered[len(ordered) // 4]
    q3 = ordered[(len(ordered) * 3) // 4]
    iqr = q3 - q1
    if iqr <= 0:
        return []
    low = q1 - 1.5 * iqr
    high = q3 + 1.5 * iqr
    return [value for value in numbers if value < low or value > high]


def _to_float(value: str) -> float | None:
    text = str(value or "").strip().replace(",", "")
    if not text:
        return None
    try:
        if text.endswith("%"):
            return float(text[:-1]) / 100
        return float(text)
    except ValueError:
        return None


def _table_to_csv_bytes(headers: list[str], rows: list[list[Any]]) -> bytes:
    out = io.StringIO()
    writer = csv.writer(out, lineterminator="\n")
    writer.writerow(headers)
    writer.writerows(rows)
    return out.getvalue().encode("utf-8")


def _run_ingest_if_present(result: dict[str, Any]) -> None:
    job_id = result.get("ingest_job_id")
    workspace_id = result.get("workspace_id")
    if job_id and workspace_id:
        run_workspace_ingest_job(str(workspace_id), str(job_id))


def _safe_filename(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9_.-]+", "-", str(value or "import")).strip("-")
    return cleaned[:90] or "import"


def _clean_sql_payload(body: dict[str, Any]) -> dict[str, Any]:
    connection_string = str(body.get("connection_string") or "").strip()
    if connection_string:
        parsed = _parse_sql_connection_string(connection_string)
        if parsed:
            return parsed
        return {"connection_string": connection_string}
    server = str(body.get("server") or "").strip()
    database = str(body.get("database") or "").strip()
    username = str(body.get("username") or body.get("user") or "").strip()
    password = str(body.get("password") or "")
    if not server or not database:
        raise ValueError("server and database are required")
    if not username and not str(body.get("entra_connection_string") or "").strip():
        raise ValueError("username/password or connection_string is required")
    return {"server": server, "database": database, "username": username, "password": password}


def _sql_tables(payload: dict[str, Any]) -> list[dict[str, Any]]:
    with closing(_sql_connection(payload)) as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT TABLE_SCHEMA, TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_TYPE='BASE TABLE' ORDER BY TABLE_SCHEMA, TABLE_NAME")
        rows = cursor.fetchall()
        return [{"schema": str(row[0]), "name": str(row[1]), "id": f"{row[0]}.{row[1]}"} for row in rows]


def _sql_preview(payload: dict[str, Any], table: str, limit: int) -> dict[str, Any]:
    tables = {item["id"] for item in _sql_tables(payload)}
    if table not in tables:
        raise ValueError("Unknown table")
    query = f"SELECT TOP {int(limit)} * FROM {_quote_table(table)}"
    with closing(_sql_connection(payload)) as conn:
        cursor = conn.cursor()
        cursor.execute(query)
        columns = [item[0] for item in cursor.description or []]
        rows = [[_cell(value) for value in row] for row in cursor.fetchall()]
    return {"kind": "table", "columns": [{"name": name} for name in columns], "rows": rows, "total_rows": len(rows), "total_cols": len(columns), "source": {"kind": "sql", "table": table}}


def _sql_connection(payload: dict[str, Any]) -> Any:
    driver_payload = payload
    if payload.get("connection_string"):
        driver_payload = _parse_sql_connection_string(str(payload["connection_string"])) or payload
    try:
        import pymssql  # type: ignore

        if driver_payload.get("connection_string"):
            raise RuntimeError("This SQL connection string requires pyodbc, but the backend image only has pymssql.")
        server, port = _split_sql_server(str(driver_payload["server"]))
        options: dict[str, Any] = {
            "server": server,
            "user": driver_payload.get("username") or None,
            "password": driver_payload.get("password") or None,
            "database": driver_payload["database"],
            "login_timeout": int(driver_payload.get("login_timeout") or 8),
            "timeout": int(driver_payload.get("timeout") or 15),
        }
        if port:
            options["port"] = port
        return pymssql.connect(
            **options,
        )
    except ImportError:
        try:
            import pyodbc  # type: ignore
        except ImportError as exc:
            raise RuntimeError("SQL driver is not installed. Install pymssql or pyodbc in the backend image.") from exc
        if payload.get("connection_string"):
            connection_string = payload["connection_string"]
        else:
            connection_string = (
                "DRIVER={ODBC Driver 18 for SQL Server};"
                f"SERVER={driver_payload['server']};DATABASE={driver_payload['database']};"
                f"UID={driver_payload.get('username')};PWD={driver_payload.get('password')};"
                "Encrypt=yes;TrustServerCertificate=no;Connection Timeout=5;"
            )
        return pyodbc.connect(connection_string, timeout=8)


def _parse_sql_connection_string(connection_string: str) -> dict[str, Any] | None:
    values: dict[str, str] = {}
    for part in str(connection_string or "").split(";"):
        if not part.strip() or "=" not in part:
            continue
        key, value = part.split("=", 1)
        normalized = re.sub(r"[\s_-]+", "", key.strip().lower())
        values[normalized] = value.strip().strip('"')
    server = values.get("server") or values.get("datasource") or values.get("address") or values.get("addr") or values.get("networkaddress")
    database = values.get("database") or values.get("initialcatalog")
    username = values.get("userid") or values.get("uid") or values.get("user")
    password = values.get("password") or values.get("pwd")
    if not server or not database or not username or password is None:
        return None
    result: dict[str, Any] = {
        "server": server,
        "database": database,
        "username": username,
        "password": password,
    }
    timeout = values.get("connectiontimeout") or values.get("connecttimeout") or values.get("timeout")
    if timeout:
        try:
            result["login_timeout"] = max(1, min(30, int(float(timeout))))
        except ValueError:
            pass
    return result


def _split_sql_server(server: str) -> tuple[str, int | None]:
    value = str(server or "").strip()
    if value.lower().startswith("tcp:"):
        value = value[4:]
    if "," in value:
        host, raw_port = value.rsplit(",", 1)
        try:
            return host.strip(), int(raw_port.strip())
        except ValueError:
            return value, None
    return value, None


def _quote_table(table: str) -> str:
    parts = table.split(".")
    if len(parts) != 2 or not all(re.match(r"^[A-Za-z0-9_@$# -]+$", part) for part in parts):
        raise ValueError("Invalid table name")
    return ".".join(f"[{part.replace(']', ']]')}]" for part in parts)


def _clean_blob_payload(body: dict[str, Any]) -> dict[str, Any]:
    connection_string = str(body.get("connection_string") or "").strip()
    if connection_string:
        return {"connection_string": connection_string}
    account = str(body.get("account") or "").strip()
    sas = str(body.get("sas") or "").strip().lstrip("?")
    if not account or not sas:
        raise ValueError("connection_string or account+sas is required")
    if not re.match(r"^[a-z0-9]{3,24}$", account):
        raise ValueError("Invalid storage account name")
    return {"account": account, "sas": sas}


def _blob_service(payload: dict[str, Any]) -> Any:
    from azure.storage.blob import BlobServiceClient

    if payload.get("connection_string"):
        return BlobServiceClient.from_connection_string(payload["connection_string"])
    return BlobServiceClient(account_url=f"https://{payload['account']}.blob.core.windows.net", credential=payload["sas"])


def _blob_containers(payload: dict[str, Any]) -> list[dict[str, Any]]:
    service = _blob_service(payload)
    containers = []
    for container in service.list_containers():
        containers.append({"name": container["name"], "updated_at": _iso(container.get("last_modified"))})
        if len(containers) >= 500:
            break
    return containers


def _type_from_name(name: str, content_type: str | None = None) -> str:
    suffix = Path(str(name)).suffix.lower().lstrip(".")
    if suffix:
        return "xlsx" if suffix == "xlsm" else suffix
    content = (content_type or "").lower()
    if "csv" in content:
        return "csv"
    if "spreadsheet" in content or "excel" in content:
        return "xlsx"
    if "markdown" in content or "text/plain" in content:
        return "md"
    return "unknown"


class _ConnectorVault:
    def __init__(self) -> None:
        self._store: dict[str, dict[str, Any]] = {}
        self._fernet = self._build_fernet()

    def put(self, kind: str, workspace_id: str, payload: dict[str, Any]) -> str:
        self._purge()
        connection_id = f"{kind}_{uuid.uuid4().hex[:16]}"
        expires_at = time.time() + CONNECTOR_TTL_SECONDS
        self._store[connection_id] = {
            "kind": kind,
            "workspace_id": workspace_id,
            "expires_at": expires_at,
            "payload": self._encrypt(payload),
        }
        return connection_id

    def get(self, connection_id: str, kind: str, workspace_id: str) -> dict[str, Any]:
        self._purge()
        item = self._store.get(str(connection_id or ""))
        if not item or item.get("kind") != kind or item.get("workspace_id") != workspace_id:
            raise ValueError("Connector session not found or expired")
        return self._decrypt(bytes(item["payload"]))

    def delete(self, connection_id: str, *, kind: str | None = None, workspace_id: str | None = None) -> bool:
        self._purge()
        key = str(connection_id or "")
        item = self._store.get(key)
        if not item:
            return False
        if kind and item.get("kind") != kind:
            return False
        if workspace_id and item.get("workspace_id") != workspace_id:
            return False
        self._store.pop(key, None)
        return True

    def status(self, connection_id: str, kind: str, workspace_id: str) -> dict[str, Any]:
        self._purge()
        key = str(connection_id or "")
        item = self._store.get(key)
        if not item or item.get("kind") != kind or item.get("workspace_id") != workspace_id:
            raise ValueError("Connector session not found or expired")
        return {
            "workspace_id": workspace_id,
            "kind": kind,
            "connection_id": key,
            "status": "connected",
            "expires_at": self.expires_at(key),
        }

    def expires_at(self, connection_id: str) -> str | None:
        item = self._store.get(connection_id)
        return _iso(datetime.fromtimestamp(float(item["expires_at"]), tz=timezone.utc)) if item else None

    def _purge(self) -> None:
        now = time.time()
        for key in [key for key, value in self._store.items() if float(value.get("expires_at") or 0) <= now]:
            self._store.pop(key, None)

    def _build_fernet(self) -> Any:
        try:
            from cryptography.fernet import Fernet
        except ImportError as exc:
            raise RuntimeError("cryptography is required for connector credential encryption") from exc
        key = os.environ.get("DF_CONNECTOR_CREDENTIAL_KEY")
        if key:
            raw = key.encode("utf-8")
            if len(raw) != 44:
                raw = base64.urlsafe_b64encode(hashlib.sha256(raw).digest())
            return Fernet(raw)
        return Fernet(Fernet.generate_key())

    def _encrypt(self, payload: dict[str, Any]) -> bytes:
        return self._fernet.encrypt(json.dumps(payload, ensure_ascii=False).encode("utf-8"))

    def _decrypt(self, payload: bytes) -> dict[str, Any]:
        data = json.loads(self._fernet.decrypt(payload).decode("utf-8"))
        if not isinstance(data, dict):
            raise ValueError("Invalid connector payload")
        return data


_CONNECTORS = _ConnectorVault()


def _raise_sql_connector_error(exc: BaseException, *, status_code: int = 400) -> None:
    safe = _safe_connector_error(exc)
    print(f"sql connector failed: {safe}", flush=True)
    lowered = safe.lower()
    if "driver is not installed" in lowered or "pymssql" in lowered and "pyodbc" in lowered:
        message = "数据库连接组件不可用，请联系管理员检查后端镜像里的 SQL 驱动。"
        hint = "需要安装 pymssql 或 pyodbc，并确认容器能加载对应驱动。"
    elif any(term in lowered for term in ("login failed", "authentication", "password", "18456")):
        message = "无法登录数据库，请检查账号、密码或身份连接串是否正确。"
        hint = "建议先用只读账号验证；不要在截图或聊天里暴露密码。"
    elif any(term in lowered for term in ("timeout", "timed out", "refused", "network", "db-lib", "server is unavailable", "could not open")):
        message = "无法连接到数据库，请检查服务器地址、数据库名和防火墙是否允许当前后端访问。"
        hint = "Azure SQL 常见原因是未放行客户端 IP、服务器名写错、端口 1433 不通或数据库暂不可用。"
    else:
        message = "数据库连接失败，请检查连接配置后重试。"
        hint = "请确认服务器、数据库、账号权限和网络访问策略均正确。"
    raise HTTPException(
        status_code=status_code,
        detail={"message": message, "hint": hint, "error_type": "sql_connection_error"},
    ) from exc


def _safe_connector_error(exc: BaseException) -> str:
    text = f"{type(exc).__name__}: {exc}"
    text = re.sub(r"(?i)(password|pwd|sig|accountkey|sharedaccesssignature)=([^;\\s]+)", r"\1=***", text)
    return text[:500]


def _parse_sse_frame(raw: str) -> list[tuple[str, Any]]:
    frames: list[tuple[str, Any]] = []
    for block in str(raw or "").split("\n\n"):
        event = ""
        data_lines: list[str] = []
        for line in block.splitlines():
            if line.startswith("event:"):
                event = line.removeprefix("event:").strip()
            elif line.startswith("data:"):
                data_lines.append(line.removeprefix("data:").strip())
        if not event:
            continue
        data_text = "\n".join(data_lines)
        try:
            data: Any = json.loads(data_text) if data_text else {}
        except json.JSONDecodeError:
            data = data_text
        frames.append((event, data))
    return frames


def _compact_event_data(data: Any) -> Any:
    if not isinstance(data, dict):
        return data
    compact = dict(data)
    if "delta" in compact:
        compact["delta"] = str(compact["delta"])[:120]
    if isinstance(compact.get("artifact"), dict):
        compact["artifact"] = {
            "workspace_id": compact["artifact"].get("workspace_id"),
            "conversation_id": compact["artifact"].get("conversation_id"),
            "keys": sorted(str(key) for key in compact["artifact"].keys())[:30],
        }
    return compact


def _clamp_int(value: Any, low: int, high: int, default: int | None) -> int:
    if value is None:
        if default is None:
            raise ValueError("integer value is required")
        return default
    try:
        number = int(value)
    except (TypeError, ValueError) as exc:
        raise ValueError("invalid integer value") from exc
    if number < low or number > high:
        raise ValueError(f"integer value must be between {low} and {high}")
    return number


def _iso(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.astimezone(timezone.utc).isoformat()
    return str(value)


def _utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()
