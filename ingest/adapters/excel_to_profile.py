from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .tabular_profile import clean_cell, infer_table_profile, normalize_headers


def excel_to_profile(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    tables: list[dict[str, Any]] = []
    try:
        for sheet in workbook.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                tables.append(infer_table_profile(sheet.title, []))
                continue
            width = max(len(row) for row in rows)
            headers = normalize_headers(list(rows[0]), width)
            records = []
            for raw in rows[1:]:
                if not any(clean_cell(value) for value in raw):
                    continue
                records.append({headers[idx]: clean_cell(raw[idx] if idx < len(raw) else "") for idx in range(width)})
            tables.append(infer_table_profile(sheet.title, records))
    finally:
        workbook.close()
    return tables
