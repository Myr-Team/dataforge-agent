from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def _safe_id(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9_-]+", "-", value).strip("-") or "excel"


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def excel_to_records(path: Path, rel_path: str, workspace_id: str) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    records: list[dict[str, Any]] = []
    for sheet in workbook.worksheets:
        rows = sheet.iter_rows(values_only=True)
        try:
            headers = [_cell_text(value) or f"column_{idx + 1}" for idx, value in enumerate(next(rows))]
        except StopIteration:
            continue
        for row_number, values in enumerate(rows, start=2):
            pairs = [
                (headers[idx] if idx < len(headers) else f"column_{idx + 1}", _cell_text(value))
                for idx, value in enumerate(values)
                if _cell_text(value)
            ]
            if not pairs:
                continue
            content = "; ".join(f"{key}: {value}" for key, value in pairs)
            sheet_id = _safe_id(sheet.title)
            file_id = _safe_id(path.stem)
            chunk_id = f"{file_id}-{sheet_id}-row-{row_number}"
            records.append(
                {
                    "@search.action": "mergeOrUpload",
                    "id": f"{workspace_id}-{chunk_id}",
                    "workspace_id": workspace_id,
                    "title": f"{path.stem} / {sheet.title} row {row_number}",
                    "content": content,
                    "source_file": rel_path,
                    "chunk_id": chunk_id,
                    "document_type": "excel",
                    "language": "zh-Hans" if any("\u4e00" <= char <= "\u9fff" for char in content) else "en",
                    "sheet": sheet.title,
                    "row": str(row_number),
                }
            )
    workbook.close()
    return records
